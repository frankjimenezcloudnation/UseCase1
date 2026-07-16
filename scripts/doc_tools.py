#!/usr/bin/env python3
"""Geeft Claude Code agents teksttoegang tot binaire referentiedocumenten.

Subcommando's:
  ontology-search <term>        Zoekt case-insensitive in ontologie-XLSX.
                                Voorbeeld: python doc_tools.py ontology-search "persoon" --limit 10
  ontology-list [sheet]         Toont alle entries van een ontologie-sheet.
                                Voorbeeld: python doc_tools.py ontology-list CLASS
  docx-text <pad>               Extraheert leesbare tekst uit een .docx-bestand.
                                Voorbeeld: python doc_tools.py docx-text "AnalyseQwik_FPR_202508.docx"
  xlsx-dump <pad> [sheet]       Toont ruwe data uit een .xlsx-bestand.
                                Voorbeeld: python doc_tools.py xlsx-dump OntologySnapshot.xlsx CLASS

Vereisten: openpyxl, python-docx (gebruik backend/.venv/bin/python om uit te voeren).
"""

import argparse
import signal
import sys
from pathlib import Path

# Nette afsluiting bij pipen naar bv. `head`
try:
    signal.signal(signal.SIGPIPE, signal.SIG_DFL)
except (AttributeError, ValueError):
    pass

REPO_ROOT = Path(__file__).resolve().parent.parent

_IMPORT_FOUT = "FOUT: draai dit script via backend/.venv/bin/python (openpyxl/python-docx vereist)"


def _resolve_path(path_str: str) -> Path:
    p = Path(path_str)
    if not p.is_absolute():
        p = REPO_ROOT / p
    return p.resolve()


def _check_file(path: Path) -> None:
    if not path.is_file():
        print(f"FOUT: bestand niet gevonden: {path}", file=sys.stderr)
        sys.exit(2)


def _truncate(text: str, max_len: int = 160) -> str:
    if len(text) > max_len:
        return text[:max_len] + "…"
    return text


def _col_indices(header_row, *names):
    indices = {}
    for i, cell in enumerate(header_row):
        val = str(cell.value).strip() if cell.value is not None else ""
        for name in names:
            if val.lower() == name.lower():
                indices[name] = i
    return indices


def _sheet_has_syntax(sheet_name: str) -> bool:
    return sheet_name.upper() in ("CONDITION", "RULE")


def _import_openpyxl():
    try:
        import openpyxl
    except ImportError:
        print(_IMPORT_FOUT, file=sys.stderr)
        sys.exit(3)
    return openpyxl


def cmd_ontology_search(args):
    openpyxl = _import_openpyxl()
    file_path = _resolve_path(args.file)
    _check_file(file_path)

    target_sheets = {"SUPERCLASS", "CLASS", "SUBCLASS", "PROPERTY", "CONDITION", "RULE"}
    term_lower = args.term.lower()
    limit = args.limit
    count = 0

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() not in target_sheets:
                continue
            ws = wb[sheet_name]
            rows = ws.iter_rows()
            try:
                header = next(rows)
            except StopIteration:
                continue

            idx = _col_indices(header, "ExternalId", "Name", "Syntax", "Definition", "Clarification")
            ext_id_col = idx.get("ExternalId")
            name_or_syntax_col = idx.get("Syntax") if _sheet_has_syntax(sheet_name) else idx.get("Name")
            def_col = idx.get("Definition")
            clar_col = idx.get("Clarification")

            for row in rows:
                if count >= limit:
                    break

                cells = list(row)

                def _cell_val(col_idx):
                    if col_idx is None or col_idx >= len(cells):
                        return ""
                    v = cells[col_idx].value
                    return str(v).strip() if v is not None else ""

                name_text = _cell_val(name_or_syntax_col)
                def_text = _cell_val(def_col)
                clar_text = _cell_val(clar_col)

                combined = f"{name_text} {def_text} {clar_text}".lower()
                if term_lower not in combined:
                    continue

                ext_id = _cell_val(ext_id_col)
                print(f"{sheet_name}\t{ext_id}\t{name_text}\t{_truncate(def_text)}")
                count += 1

            if count >= limit:
                break
    finally:
        wb.close()

    if count == 0:
        print(f"# 0 resultaten voor '{args.term}' — markeer de term als ontologie-afwijking")
    else:
        print(f"# {count} resultaten (limiet {limit})")


def cmd_ontology_list(args):
    openpyxl = _import_openpyxl()
    file_path = _resolve_path(args.file)
    _check_file(file_path)

    sheet_name_input = args.sheet.upper()

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        matches = [s for s in wb.sheetnames if s.upper() == sheet_name_input]
        if not matches:
            print(
                f"FOUT: sheet '{args.sheet}' niet gevonden. Beschikbaar: {', '.join(wb.sheetnames)}",
                file=sys.stderr,
            )
            sys.exit(2)

        sn = matches[0]
        ws = wb[sn]
        rows = ws.iter_rows()
        try:
            header = next(rows)
        except StopIteration:
            return

        idx = _col_indices(header, "Name", "Syntax", "Definition")
        name_or_syntax_col = idx.get("Syntax") if _sheet_has_syntax(sn) else idx.get("Name")
        def_col = idx.get("Definition")

        for row in rows:
            cells = list(row)

            def _cell_val(col_idx):
                if col_idx is None or col_idx >= len(cells):
                    return ""
                v = cells[col_idx].value
                return str(v).strip() if v is not None else ""

            name_text = _cell_val(name_or_syntax_col)
            def_text = _cell_val(def_col)
            print(f"{name_text}\t{_truncate(def_text)}")
    finally:
        wb.close()


def cmd_docx_text(args):
    try:
        from docx import Document
    except ImportError:
        print(_IMPORT_FOUT, file=sys.stderr)
        sys.exit(3)

    file_path = _resolve_path(args.path)
    _check_file(file_path)

    doc = Document(str(file_path))

    for elem in doc.element.body:
        tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
        if tag == "p":
            text = ""
            for child in elem.iter():
                if child.tag.endswith("}t") and child.text:
                    text += child.text
            stripped = text.strip()
            if stripped:
                print(stripped)
        elif tag == "tbl":
            print("--- tabel ---")
            for row in elem.iter():
                if row.tag.endswith("}tr"):
                    cells = []
                    for cell in row.iter():
                        if cell.tag.endswith("}tc"):
                            cell_text = ""
                            for p in cell.iter():
                                if p.tag.endswith("}t") and p.text:
                                    cell_text += p.text
                            cells.append(cell_text.strip())
                    if cells:
                        print("\t".join(cells))
            print("--- einde tabel ---")


def cmd_xlsx_dump(args):
    openpyxl = _import_openpyxl()
    file_path = _resolve_path(args.path)
    _check_file(file_path)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(
                f"FOUT: sheet '{args.sheet}' niet gevonden. Beschikbaar: {', '.join(wb.sheetnames)}",
                file=sys.stderr,
            )
            wb.close()
            sys.exit(2)
        sheets_to_dump = [args.sheet]
    else:
        print(f"# sheets: {', '.join(wb.sheetnames)}")
        sheets_to_dump = list(wb.sheetnames)

    first = True
    try:
        for sn in sheets_to_dump:
            if not first:
                print()
            first = False
            print(f"## {sn}")
            ws = wb[sn]
            for row in ws.iter_rows():
                vals = []
                for cell in row:
                    v = cell.value
                    vals.append("" if v is None else _truncate(str(v), 200))
                print("\t".join(vals))
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command", required=True)

    p_s = sub.add_parser("ontology-search", help="Zoek in ontologie-XLSX")
    p_s.add_argument("term", help="Zoekterm (case-insensitive)")
    p_s.add_argument("--file", default="OntologySnapshot.xlsx", help="Pad naar XLSX (default: OntologySnapshot.xlsx)")
    p_s.add_argument("--limit", type=int, default=50, help="Maximum aantal resultaten (default: 50)")
    p_s.set_defaults(func=cmd_ontology_search)

    p_l = sub.add_parser("ontology-list", help="Toon alle entries van een ontologie-sheet")
    p_l.add_argument("sheet", nargs="?", default="CLASS", help="Sheetnaam (default: CLASS)")
    p_l.add_argument("--file", default="OntologySnapshot.xlsx", help="Pad naar XLSX (default: OntologySnapshot.xlsx)")
    p_l.set_defaults(func=cmd_ontology_list)

    p_d = sub.add_parser("docx-text", help="Extraheer tekst uit .docx")
    p_d.add_argument("path", help="Pad naar .docx-bestand")
    p_d.set_defaults(func=cmd_docx_text)

    p_x = sub.add_parser("xlsx-dump", help="Toon ruwe data uit .xlsx")
    p_x.add_argument("path", help="Pad naar .xlsx-bestand")
    p_x.add_argument("sheet", nargs="?", help="Sheetnaam (optioneel)")
    p_x.set_defaults(func=cmd_xlsx_dump)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
