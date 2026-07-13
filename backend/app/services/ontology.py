"""Deterministic retrieval over the OntologySnapshot — the 'single source of truth'.

The ontology is a large structured data model (thousands of properties/classes with
definitions and domain values). We do NOT dump all of it into the prompt, and we do NOT
use vector embeddings (whose similarity ranking would be non-deterministic). Instead we do
transparent, reproducible keyword retrieval: per WTP theme we score every ontology concept
by how many theme keywords it contains and keep the top matches.

Two products come out of this module:
  • `full_text(rows)`   → the complete, human-readable rendering of the ontology, used as the
                          verification corpus so any ontology quote can be checked.
  • `retrieve(rows, k)` → a compact per-theme block injected into the prompt so the standard
                          side is grounded in the actual data model, not the model's memory.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

# The FIXED set of themes analysed every run — this is what makes the number of findings
# stable and reproducible (no more "8 vs 10 onderdelen"). Keep in sync with the frontend
# WTP_THEMES list.
CANONICAL_THEMES: list[str] = [
    "Opbouwsystematiek",
    "Partnerpensioen",
    "Wezenpensioen",
    "Arbeidsongeschiktheid (premievrijstelling)",
    "Indexatie/Toeslagen",
    "Compensatie afschaffing doorsneesystematiek",
    "Beleggingsbeleid/Lifecycle",
    "Uitkeringsfase / Collectief Variabel Pensioen",
    "Risicodelingsreserve",
    "Invaren",
    "Bijsparen",
]

# Keywords used to retrieve relevant ontology concepts per theme (deterministic scoring).
THEME_KEYWORDS: dict[str, list[str]] = {
    "Opbouwsystematiek": ["premie", "opbouw", "premiepercentage", "pensioengrondslag",
                          "franchise", "beschikbare", "kapitaal", "grondslag"],
    "Partnerpensioen": ["partnerpensioen", "partner", "nabestaande", "overlijden", "lpp"],
    "Wezenpensioen": ["wees", "wezen", "kind"],
    "Arbeidsongeschiktheid (premievrijstelling)": ["arbeidsongeschikt", "premievrijstelling",
                          "premievrij", "invaliditeit", "voortzetting"],
    "Indexatie/Toeslagen": ["toeslag", "indexatie", "aanpassing", "rendement", "spreiding",
                          "projectierendement"],
    "Compensatie afschaffing doorsneesystematiek": ["compensatie", "doorsnee",
                          "doorsneesystematiek", "depot", "staffel"],
    "Beleggingsbeleid/Lifecycle": ["belegg", "lifecycle", "life cycle", "rebalanc",
                          "risicoprofiel", "units", "beleggingsbalans"],
    "Uitkeringsfase / Collectief Variabel Pensioen": ["uitkering", "collectief variabel",
                          "cvp", "variabele uitkering", "uitkeringsfase"],
    "Risicodelingsreserve": ["risicodelingsreserve", "reserve", "solidariteit", "rdr",
                          "stabiliteit"],
    "Invaren": ["invaren", "invaar", "waardeoverdracht", "standaardmethode", "transitie"],
    "Bijsparen": ["bijspar", "vrijwillig", "extra", "upa", "bijspaarruimte"],
}

_SHEETS = ("PROPERTY", "CLASS", "SUBCLASS")


@dataclass
class Concept:
    external_id: str
    name: str
    definition: str
    clarification: str
    domain_values: str
    category: str
    sheet: str
    text: str  # normalised searchable blob (name + definition + clarification + domain + category)


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "n/d", "nan") else s


def load_concepts(path: Path) -> list[Concept]:
    """Read ExternalId/Name/Definition/Clarification/DomainValue/Category from the main sheets.

    Definition and Clarification are kept SEPARATE (the frontend shows both verbatim so experts
    can verify the standard side directly against the single source of truth)."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    concepts: list[Concept] = []
    for sheet in _SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else "" for c in rows[0]]

        def col(name: str) -> int | None:
            return header.index(name) if name in header else None

        i_id, i_name = col("ExternalId"), col("Name")
        i_def, i_clar = col("Definition"), col("Clarification")
        i_dom, i_cat = col("DomainValue"), col("Category")

        def get(r, i: int | None) -> str:
            return _clean(r[i]) if i is not None and i < len(r) else ""

        for r in rows[1:]:
            name = get(r, i_name)
            if not name:
                continue
            definition = get(r, i_def)
            clarification = get(r, i_clar)
            dom = get(r, i_dom)
            cat = get(r, i_cat)
            blob = " ".join((name, definition, clarification, dom, cat)).lower()
            concepts.append(Concept(
                external_id=get(r, i_id), name=name, definition=definition,
                clarification=clarification, domain_values=dom, category=cat,
                sheet=sheet, text=blob,
            ))
    wb.close()
    return concepts


def _render(c: Concept, max_def: int | None = None) -> str:
    def cap(s: str) -> str:
        return s[:max_def].rstrip() + "…" if max_def and len(s) > max_def else s

    parts = [c.name.replace("_", " ")]
    if c.definition:
        parts.append(f": {cap(c.definition)}")
    if c.clarification:
        parts.append(f" — toelichting: {cap(c.clarification)}")
    if c.domain_values:
        parts.append(f" [waarden: {c.domain_values}]")
    if c.category:
        parts.append(f" (categorie: {c.category})")
    return "".join(parts)


def full_text(concepts: list[Concept], max_chars: int) -> str:
    """Complete rendering used as the verification corpus for ontology citations."""
    out: list[str] = []
    total = 0
    for c in concepts:
        line = "• " + _render(c)
        out.append(line)
        total += len(line)
        if total >= max_chars:
            out.append("[... ontology afgekapt voor lengte ...]")
            break
    return "\n".join(out)


def _top_for_theme(concepts: list[Concept], theme: str, k: int) -> list[Concept]:
    """Deterministic keyword ranking of concepts for one theme (stable tie-break on name)."""
    kws = THEME_KEYWORDS.get(theme, [])
    scored: list[tuple[int, str, Concept]] = []
    for c in concepts:
        score = sum(c.text.count(kw) for kw in kws)
        if score > 0:
            scored.append((score, c.name, c))
    scored.sort(key=lambda t: (-t[0], t[1]))
    return [c for _, _, c in scored[:k]]


def retrieve(concepts: list[Concept], per_theme: int = 8, max_def: int = 300) -> str:
    """Deterministic per-theme retrieval block for the PROMPT (capped definitions)."""
    blocks: list[str] = []
    for theme in CANONICAL_THEMES:
        top = _top_for_theme(concepts, theme, per_theme)
        if not top:
            continue
        lines = "\n".join("  • " + _render(c, max_def=max_def) for c in top)
        blocks.append(f"## Thema: {theme}\n{lines}")
    return "\n\n".join(blocks)


def retrieve_one(concepts: list[Concept], theme: str, per_theme: int = 8, max_def: int = 300) -> str:
    """Prompt block of the retrieved ontology concepts for a SINGLE theme (for per-theme calls)."""
    top = _top_for_theme(concepts, theme, per_theme)
    if not top:
        return "(geen specifieke ontology-concepten gevonden voor dit thema)"
    return "\n".join("• " + _render(c, max_def=max_def) for c in top)


def retrieve_structured(concepts: list[Concept], per_theme: int = 6) -> dict[str, list[Concept]]:
    """Per-theme retrieved concepts as structured objects, for VISIBLE display in the UI.

    These are attached verbatim (name/definition/clarification straight from the snapshot) so
    experts see exactly which ontology rows back the standard side of each theme."""
    return {theme: _top_for_theme(concepts, theme, per_theme) for theme in CANONICAL_THEMES}
